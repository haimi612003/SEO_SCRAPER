"""Xuất kết quả ra Excel — mỗi keyword 1 sheet."""
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

EXCEL_CELL_LIMIT = 32767
HEADERS = ["keyword", "rank", "url", "title", "content"]
COL_WIDTHS = [25, 6, 50, 50, 100]


def _safe_sheet_name(keyword: str, used: set) -> str:
    """Excel: tên sheet tối đa 31 ký tự, cấm []:*?/\\ và không được trùng."""
    name = re.sub(r"[\[\]:*?/\\]", " ", keyword).strip() or "sheet"
    name = name[:31]
    base, i = name, 2
    while name.lower() in used:
        suffix = f"~{i}"
        name = base[: 31 - len(suffix)] + suffix
        i += 1
    used.add(name.lower())
    return name


def _safe_cell(value: str) -> str:
    if value is None:
        return ""
    value = str(value)
    # loại ký tự điều khiển không hợp lệ trong xlsx
    value = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", value)
    if len(value) > EXCEL_CELL_LIMIT:
        note = "\n[TRUNCATED - nội dung vượt giới hạn 32,767 ký tự của Excel]"
        value = value[: EXCEL_CELL_LIMIT - len(note)] + note
    return value


def export_excel(data: dict[str, list[dict]], output_path: str) -> None:
    """data: {keyword: [{rank, url, title, content}, ...]}"""
    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="4472C4")
    used_names: set = set()

    for keyword, rows in data.items():
        ws = wb.create_sheet(_safe_sheet_name(keyword, used_names))

        for col, header in enumerate(HEADERS, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            ws.column_dimensions[get_column_letter(col)].width = COL_WIDTHS[col - 1]
        ws.freeze_panes = "A2"

        for r, row in enumerate(rows, start=2):
            ws.cell(row=r, column=1, value=_safe_cell(keyword))
            ws.cell(row=r, column=2, value=row.get("rank"))
            url_cell = ws.cell(row=r, column=3, value=_safe_cell(row.get("url")))
            url = row.get("url") or ""
            if url.startswith("http"):
                url_cell.hyperlink = url
                url_cell.font = Font(color="0563C1", underline="single")
            ws.cell(row=r, column=4, value=_safe_cell(row.get("title")))
            content_cell = ws.cell(row=r, column=5, value=_safe_cell(row.get("content")))
            content_cell.alignment = Alignment(wrap_text=True, vertical="top")

    if not wb.sheetnames:
        wb.create_sheet("empty")
    wb.save(output_path)
